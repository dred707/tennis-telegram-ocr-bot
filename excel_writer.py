from openpyxl import load_workbook
from io import BytesIO

def fill_excel_template(entries, filename="output.xlsx"):
    template_path = "template.xlsx"
    wb = load_workbook(template_path)
    ws = wb["main"]

    row = 2
    for entry in entries:
        table = int("".join(filter(str.isdigit, entry["Стіл"])))
        ws[f"I{row}"] = table
        ws[f"J{row}"] = entry["З"]
        ws[f"K{row}"] = entry["По"]
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output